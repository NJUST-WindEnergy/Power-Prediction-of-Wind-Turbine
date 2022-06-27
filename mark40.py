# -*- coding: utf-8 -*-
"""
Created on Wed Jan 19 21:35:40 2022

@author: LiuJingAAA
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn import metrics
# import tensorflow as tf
import tensorflow.compat.v1 as tf

tf.disable_v2_behavior()
import math
from sklearn.metrics import mean_absolute_error
import openpyxl as op


#京能风电场风轮半径38.5 Cp0.5 轴向诱导因子0.190983  响水风电场风轮半径59 Cp0.55 轴向诱导因子0.24
#京能风电场切入风速3 额定风速10   响水风电场切入风速3 额定风速8.4
    
# 权值初始化
def weight_variable(shape):
    # 用正态分布来初始化权值
  #  initial = tf.truncated_normal(shape, stddev=1)
    initial = tf.truncated_normal(shape, stddev=1, seed=1)
    return tf.Variable(initial,name='W')
 
def bias_variable(shape):
    initial = tf.constant(1.0, shape=shape)
    return tf.Variable(initial,name='b')

def add_layer(layername,inputs, in_size, out_size, activation_function=None):
    # add one more layer and return the output of this layer
    Weights = weight_variable([in_size,out_size])
    biases = bias_variable([out_size])
    Wx_plus_b = tf.add(tf.matmul(inputs, Weights), biases)
    if activation_function is None:
        outputs = Wx_plus_b
    else:
        outputs = activation_function(Wx_plus_b)
    return outputs
#tensorflow所有的计算都是计算图上的一个节点，定义好的运算是通过会话（session)来执行的，
#使用tf.concat()函数后也需要在session中执行，但是我只是对两个常量数组进行拼接，
#此时正确的函数应该是np.concatenate()，它可以直接使用
def add_layer_w(layername,x1,inputs, in_size, out_size, activation_function=None):
    x_input = x1
    inputsA = tf.concat([x_input, inputs], axis=1)
    Weights = weight_variable([in_size,out_size])
    biases = bias_variable([out_size])
    Wx_plus_b = tf.add(tf.matmul(inputsA, Weights), biases)
    if activation_function is None:
        outputs = Wx_plus_b
    else:
        outputs = activation_function(Wx_plus_b)
    return outputs



#########################数据集准备############################################

def Data(rate):
###############训练测试集##############################
    train_data=pd.read_excel(r"E:/柳靖研究生期间工作/期刊相关文件/最终使用数据\train.xlsx")
    train_data=train_data.values.astype(np.float32)
    test_data=pd.read_excel(r"E:/柳靖研究生期间工作/期刊相关文件/最终使用数据\test.xlsx")
    test_data=test_data.values.astype(np.float32)
    vaild_data=pd.read_excel(r"E:/柳靖研究生期间工作/期刊相关文件/最终使用数据\gen.xlsx")
    vaild_data=vaild_data.values.astype(np.float32)
    feature=train_data[:,:6]
    output=train_data[:,6:7]
    test_feature=test_data[:,:6]
    test_output=test_data[:,6:7]
    vaild_feature=vaild_data[:,:6]
    vaild_output=vaild_data[:,6:7]
    return feature,output,test_feature,test_output,vaild_feature,vaild_output

#########################物理尾流模型类#########################################
def D2Jensen(Tf,To,Tef,Teo,Vf,Vo):
    test_output=pd.read_excel(r"E:/柳靖研究生期间工作/期刊相关文件/最终使用数据\test.xlsx")
    test_output=test_output.values.astype(np.float32)
    vaild_output=pd.read_excel(r"E:/柳靖研究生期间工作/期刊相关文件/最终使用数据\gen.xlsx")
    vaild_output=vaild_output.values.astype(np.float32)
    test_output=test_output[:,7:8]
    vaild_output=vaild_output[:,7:8]
    y_predict=test_output
    y_true=np.array(Teo)
    y_pred=np.array(y_predict)
    zero=np.zeros(shape=(len(y_true),1))
    print("2DJensen模型预测结果")
    y1=y_pred
    a=[]
    #RMSE
    print(np.sqrt(metrics.mean_squared_error(y_true, y_pred))) # 2.847304489713536
    a.append(np.sqrt(metrics.mean_squared_error(y_true, y_pred)))
    # MAE
    print(metrics.mean_absolute_error(y_true, y_pred)) # 1.9285714285714286
    a.append(metrics.mean_absolute_error(y_true, y_pred))
    # MRE
    print(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero)) # 1.9285714285714286
    a.append(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero))
    #r2
    print(metrics.r2_score(y_true,y_pred))
    a.append(metrics.r2_score(y_true,y_pred))

    ########################泛化性################################################
    y_true=np.array(Vo)
    y_predict=vaild_output
    y_pred=np.array(y_predict)
    zero=np.zeros(shape=(len(y_true),1))
    print("2DJensen模型泛化性结果")
    y2=y_pred
    b=[]
    #RMSE
    print(np.sqrt(metrics.mean_squared_error(y_true, y_pred))) # 2.847304489713536
    b.append(np.sqrt(metrics.mean_squared_error(y_true, y_pred)))
    # MAE
    print(metrics.mean_absolute_error(y_true, y_pred)) # 1.9285714285714286
    b.append(metrics.mean_absolute_error(y_true, y_pred))
    # MRE
    print(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero)) # 1.9285714285714286
    b.append(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero))
    #r2
    print(metrics.r2_score(y_true,y_pred))
    b.append(metrics.r2_score(y_true,y_pred))

    return a,y1,b,y2
#########################神经网络类############################################
    

#————————————————————————神经网络—————————————————————————————————————————————
def network(Tf,To,Tef,Teo,Vf,Vo):    
    x= tf.placeholder(tf.float32, shape=(None,6))
    y= tf.placeholder(tf.float32, shape=(None,1))
    hidden_layer1=add_layer('hidden_layer1',x,6,16,activation_function=tf.nn.relu)
    hidden_layer2=add_layer('hidden_layer2',hidden_layer1,16,32,activation_function=tf.nn.relu)
    hidden_layer3=add_layer('hidden_layer3',hidden_layer2,32,8,activation_function=tf.nn.relu)
    hidden_layer4=add_layer('hidden_layer4',hidden_layer3,8,4,activation_function=tf.nn.relu)
    y_pre=add_layer('output_layer',hidden_layer4,4,1)
    loss_mse = tf.reduce_mean(tf.square(y_pre-y))
    train_step = tf.train.AdamOptimizer(0.03).minimize(loss_mse)
    #train_step = tf.train.GradientDescentOptimizer(0.03).minimize(loss_mse)AdamOptimizer
    ######################################################################
    #######################读取数据#########################################
    feature=Tf
    output=To
    test_feature=Tef
    test_output=Teo
    vaild_feature=Vf
    vaild_output=Vo
    ###########################################################################3
    ################################训练神经网络################################
    with tf.Session() as sess:
        sess.run(tf.global_variables_initializer())
        #设定训练次数
        STEPS = 30000
        trainmse=[]
        testmse=[]
        for i in range(STEPS):
            #更新参数
    #在使用改进的损失函数，用到了全部的x作为损失函数的参数时，不能使用批量训练 或者批量训练时让对应的x也批量输入 做一个placeholder
            sess.run(train_step,
                feed_dict={x: feature, y: output})
            #定期输出
            if i % 500 ==0:
                #计算损失函数
                train_mse = sess.run(
                    loss_mse, feed_dict={x: feature, y: output})
                test_mse = sess.run(
                    loss_mse, feed_dict={x: test_feature, y: test_output})
                trainmse.append(train_mse)
                testmse.append(test_mse)
                print("After %d  training step(s), trainmse on all data is %g" 
                    %(i, train_mse))
                print("testmse on all data is %g" 
                    %(test_mse))
    
    #########################################################################
    ###########################测试准确率#####################################
    #MAE MSE RMSE 校正R平方
        y_predict=sess.run(y_pre,feed_dict={x: test_feature}) 
        y_true=np.array(test_output)
        y_pred=np.array(y_predict)
        zero=np.zeros(shape=(len(y_true),1))
        print("神经网络预测结果")
        y1=y_pred
        a=[]
        #RMSE
        print(np.sqrt(metrics.mean_squared_error(y_true, y_pred))) # 2.847304489713536
        a.append(np.sqrt(metrics.mean_squared_error(y_true, y_pred)))
        # MAE
        print(metrics.mean_absolute_error(y_true, y_pred)) # 1.9285714285714286
        a.append(metrics.mean_absolute_error(y_true, y_pred))
        # MRE
        print(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero)) # 1.9285714285714286
        a.append(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero))
        #r2
        print(metrics.r2_score(y_true,y_pred))
        a.append(metrics.r2_score(y_true,y_pred))

        ########################泛化性################################################
        y_true=np.array(vaild_output)
        y_predict=sess.run(y_pre,feed_dict={x: vaild_feature}) 
        y_pred=np.array(y_predict)
        zero=np.zeros(shape=(len(y_true),1))
        print("神经网络泛化性结果")
        y2=y_pred
        b=[]
        #RMSE
        print(np.sqrt(metrics.mean_squared_error(y_true, y_pred))) # 2.847304489713536
        b.append(np.sqrt(metrics.mean_squared_error(y_true, y_pred)))
        # MAE
        print(metrics.mean_absolute_error(y_true, y_pred)) # 1.9285714285714286
        b.append(metrics.mean_absolute_error(y_true, y_pred))
        # MRE
        print(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero)) # 1.9285714285714286
        b.append(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero))
        #r2
        print(metrics.r2_score(y_true,y_pred))
        b.append(metrics.r2_score(y_true,y_pred))

    return a,y1,b,y2
  
#————————————————————————————前端Jensen模型+神经网络——————————————————————————
    
def Jensennetwork(Tf,To,Tef,Teo,Vf,Vo):    
    x= tf.placeholder(tf.float32, shape=(None,12))
    y= tf.placeholder(tf.float32, shape=(None,1))
    hidden_layer1=add_layer('hidden_layer1',x,12,32,activation_function=tf.nn.relu)
    hidden_layer2=add_layer('hidden_layer2',hidden_layer1,32,16,activation_function=tf.nn.relu)
    hidden_layer3=add_layer('hidden_layer3',hidden_layer2,16,8,activation_function=tf.nn.relu)
    hidden_layer4=add_layer('hidden_layer4',hidden_layer3,8,4,activation_function=tf.nn.relu)
    y_pre=add_layer('output_layer',hidden_layer4,4,1)
    loss_mse = tf.reduce_mean(tf.square(y_pre-y))
    train_step = tf.train.AdamOptimizer(0.03).minimize(loss_mse)
    #train_step = tf.train.GradientDescentOptimizer(0.03).minimize(loss_mse)
    ######################################################################
    #######################读取数据#########################################
    output=To
    test_output=Teo
    vaild_output=Vo
    train_data1=pd.read_excel(r"E:/柳靖研究生期间工作/期刊相关文件/最终使用数据\train.xlsx")
    train_data1=train_data1.values.astype(np.float32)
    phy_train_feature=train_data1[:,8:14]
    test_data1=pd.read_excel(r"E:/柳靖研究生期间工作/期刊相关文件/最终使用数据\test.xlsx")
    test_data1=test_data1.values.astype(np.float32)
    phy_test_feature=test_data1[:,8:14]
    vaild_data1=pd.read_excel(r"E:/柳靖研究生期间工作/期刊相关文件/最终使用数据\gen.xlsx")
    vaild_data1=vaild_data1.values.astype(np.float32)
    phy_vaild_feature=vaild_data1[:,8:14]
    feature=np.concatenate([Tf,phy_train_feature],axis=1)
    test_feature=np.concatenate([Tef,phy_test_feature],axis=1)
    vaild_feature=np.concatenate([Vf,phy_vaild_feature],axis=1)
    ###########################################################################3
    ################################训练神经网络################################
    with tf.Session() as sess:
        sess.run(tf.global_variables_initializer())
        #设定训练次数
        STEPS = 30000
        trainmse=[]
        testmse=[]
        for i in range(STEPS):
            #更新参数
    #在使用改进的损失函数，用到了全部的x作为损失函数的参数时，不能使用批量训练 或者批量训练时让对应的x也批量输入 做一个placeholder
            sess.run(train_step,
                feed_dict={x: feature, y: output})
            #定期输出
            if i % 5 ==0:
                #计算损失函数
                train_mse = sess.run(
                    loss_mse, feed_dict={x: feature, y: output})
                test_mse = sess.run(
                    loss_mse, feed_dict={x: test_feature, y: test_output})
                trainmse.append(train_mse)
                testmse.append(test_mse)
#                print("After %d  training step(s), trainmse on all data is %g" 
#                    %(i, train_mse))
#                print("testmse on all data is %g" 
#                    %(test_mse))
    
    #########################################################################
    ###########################测试准确率#####################################
    #MAE MSE RMSE 校正R平方
        y_predict=sess.run(y_pre,feed_dict={x: test_feature}) 
        y_true=np.array(test_output)
        y_pred=np.array(y_predict)
        zero=np.zeros(shape=(len(y_true),1))
        print("2DJensen模型前端神经网络预测结果")
        y1=y_pred
        a=[]
        #RMSE
        print(np.sqrt(metrics.mean_squared_error(y_true, y_pred))) # 2.847304489713536
        a.append(np.sqrt(metrics.mean_squared_error(y_true, y_pred)))
        # MAE
        print(metrics.mean_absolute_error(y_true, y_pred)) # 1.9285714285714286
        a.append(metrics.mean_absolute_error(y_true, y_pred))
        # MRE
        print(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero)) # 1.9285714285714286
        a.append(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero))
        #r2
        print(metrics.r2_score(y_true,y_pred))
        a.append(metrics.r2_score(y_true,y_pred))
        ########################泛化性################################################
        y_true=np.array(vaild_output)
        y_predict=sess.run(y_pre,feed_dict={x: vaild_feature}) 
        y_pred=np.array(y_predict)
        zero=np.zeros(shape=(len(y_true),1))
        print("2DJensen模型前端神经网络泛化性结果")
        y2=y_pred
        b=[]
        #RMSE
        print(np.sqrt(metrics.mean_squared_error(y_true, y_pred))) # 2.847304489713536
        b.append(np.sqrt(metrics.mean_squared_error(y_true, y_pred)))
        # MAE
        print(metrics.mean_absolute_error(y_true, y_pred)) # 1.9285714285714286
        b.append(metrics.mean_absolute_error(y_true, y_pred))
        # MRE
        print(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero)) # 1.9285714285714286
        b.append(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero))
        #r2
        print(metrics.r2_score(y_true,y_pred))
        b.append(metrics.r2_score(y_true,y_pred))
    return a,y1,b,y2


#————————————————————————————神经网络+Jensen模型后端————————————————————————
def networkJensen(Tf,To,Tef,Teo,Vf,Vo):
    x= tf.placeholder(tf.float32, shape=(None,6))
    y= tf.placeholder(tf.float32, shape=(None,1))
    rule=tf.placeholder(tf.float32, shape=(None,1))
    hidden_layer1=add_layer('hidden_layer1',x,6,16,activation_function=tf.nn.relu)
    hidden_layer2=add_layer('hidden_layer2',hidden_layer1,16,32,activation_function=tf.nn.relu)
    hidden_layer3=add_layer('hidden_layer3',hidden_layer2,32,8,activation_function=tf.nn.relu)
    hidden_layer4=add_layer('hidden_layer4',hidden_layer3,8,4,activation_function=tf.nn.relu)
    y_pre=add_layer('output_layer',hidden_layer4,4,1)
    loss_mse = tf.reduce_mean(tf.square(y_pre-y)+tf.square(y_pre-rule))
    train_step = tf.train.AdamOptimizer(0.03).minimize(loss_mse)
    #train_step = tf.train.GradientDescentOptimizer(0.03).minimize(loss_mse)AdamOptimizer
    ######################################################################
    #######################读取数据#########################################
    feature=Tf
    output=To
    test_feature=Tef
    test_output=Teo
    vaild_feature=Vf
    vaild_output=Vo
    train_output1=pd.read_excel(r"E:/柳靖研究生期间工作/期刊相关文件/最终使用数据\train.xlsx")
    train_output1=train_output1.values.astype(np.float32)
    test_output1=pd.read_excel(r"E:/柳靖研究生期间工作/期刊相关文件/最终使用数据\test.xlsx")
    test_output1=test_output1.values.astype(np.float32)
    Tjensenpower=train_output1[:,7:8]
    Tejensenpower=test_output1[:,7:8]
    ###########################################################################3
    ################################训练神经网络################################
    with tf.Session() as sess:
        sess.run(tf.global_variables_initializer())
        #设定训练次数
        STEPS = 30000
        trainmse=[]
        testmse=[]
        for i in range(STEPS):
            #更新参数
    #在使用改进的损失函数，用到了全部的x作为损失函数的参数时，不能使用批量训练 或者批量训练时让对应的x也批量输入 做一个placeholder
            sess.run(train_step,
                feed_dict={x: feature,rule:Tjensenpower, y: output})
            #定期输出
            if i % 5 ==0:
                #计算损失函数
                train_mse = sess.run(
                    loss_mse, feed_dict={x: feature,rule:Tjensenpower, y: output})
                test_mse = sess.run(
                    loss_mse, feed_dict={x: test_feature,rule:Tejensenpower, y: test_output})
                trainmse.append(train_mse)
                testmse.append(test_mse)
    #            print("After %d  training step(s), trainmse on all data is %g" 
    #                %(i, train_mse))
    #            print("testmse on all data is %g" 
    #                %(test_mse))
    

    ###########################测试准确率#####################################
    #MAE MSE RMSE 校正R平方
        y_predict=sess.run(y_pre,feed_dict={x: test_feature}) 
        y_true=np.array(test_output)
        y_pred=np.array(y_predict)
        zero=np.zeros(shape=(len(y_true),1))
        print("2DJensen模型正则化神经网络预测结果")
        y1=y_pred
        a=[]
        #RMSE
        print(np.sqrt(metrics.mean_squared_error(y_true, y_pred))) # 2.847304489713536
        a.append(np.sqrt(metrics.mean_squared_error(y_true, y_pred)))
        # MAE
        print(metrics.mean_absolute_error(y_true, y_pred)) # 1.9285714285714286
        a.append(metrics.mean_absolute_error(y_true, y_pred))
        # MRE
        print(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero)) # 1.9285714285714286
        a.append(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero))
        #r2
        print(metrics.r2_score(y_true,y_pred))
        a.append(metrics.r2_score(y_true,y_pred))
        ########################泛化性################################################
        y_true=np.array(vaild_output)
        y_predict=sess.run(y_pre,feed_dict={x: vaild_feature}) 
        y_pred=np.array(y_predict)
        zero=np.zeros(shape=(len(y_true),1))
        print("2DJensen模型正则化神经网络泛化性结果")
        y2=y_pred
        b=[]
        #RMSE
        print(np.sqrt(metrics.mean_squared_error(y_true, y_pred))) # 2.847304489713536
        b.append(np.sqrt(metrics.mean_squared_error(y_true, y_pred)))
        # MAE
        print(metrics.mean_absolute_error(y_true, y_pred)) # 1.9285714285714286
        b.append(metrics.mean_absolute_error(y_true, y_pred))
        # MRE
        print(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero)) # 1.9285714285714286
        b.append(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero))
        #r2
        print(metrics.r2_score(y_true,y_pred))
        b.append(metrics.r2_score(y_true,y_pred))
    return a,y1,b,y2



#——————————————————————————————————————————权重网络————————————————————————————
def weightnetwork(Tf,To,Tef,Teo,Vf,Vo):
    x= tf.placeholder(tf.float32, shape=(None,6))
    xjensen= tf.placeholder(tf.float32, shape=(None,1))
    y= tf.placeholder(tf.float32, shape=(None,1))
    hidden_layer1=add_layer('hidden_layer1',x,6,16,activation_function=tf.nn.relu)
    hidden_layer2=add_layer('hidden_layer2',hidden_layer1,16,32,activation_function=tf.nn.relu)
    hidden_layer3=add_layer('hidden_layer3',hidden_layer2,32,8,activation_function=tf.nn.relu)
    hidden_layer4=add_layer('hidden_layer4',hidden_layer3,8,4,activation_function=tf.nn.relu)
    hidden_layer5=add_layer('hidden_layer5',hidden_layer4,4,1)
    W_layer=add_layer_w('W_layer',xjensen,hidden_layer5,2,5,activation_function=tf.nn.relu)
    hidden_layer6=add_layer('hidden_layer6',W_layer,5,3,activation_function=tf.nn.relu)
    y_pre=add_layer('output_layer',hidden_layer6,3,1)
    loss_mse = tf.reduce_mean(tf.square(y_pre-y))
    train_step = tf.train.AdamOptimizer(0.03).minimize(loss_mse)
    #train_step = tf.train.GradientDescentOptimizer(0.03).minimize(loss_mse)AdamOptimizer
    ######################################################################
    #######################读取数据#########################################
    feature=Tf
    output=To
    test_feature=Tef
    test_output=Teo
    vaild_feature=Vf
    vaild_output=Vo
    train_output1=pd.read_excel(r"E:/柳靖研究生期间工作/期刊相关文件/最终使用数据\train.xlsx")
    train_output1=train_output1.values.astype(np.float32)
    test_output1=pd.read_excel(r"E:/柳靖研究生期间工作/期刊相关文件/最终使用数据\test.xlsx")
    test_output1=test_output1.values.astype(np.float32)
    vaild_output1=pd.read_excel(r"E:/柳靖研究生期间工作/期刊相关文件/最终使用数据\gen.xlsx")
    vaild_output1=vaild_output1.values.astype(np.float32)
    TJensenpower=train_output1[:,7:8]
    TeJensenpower=test_output1[:,7:8]
    VJensenpower=vaild_output1[:,7:8]
    ###########################################################################3
    ################################训练神经网络################################
    with tf.Session() as sess:
        sess.run(tf.global_variables_initializer())
        #设定训练次数
        STEPS = 30000
        trainmse=[]
        testmse=[]
        for i in range(STEPS):
            #更新参数
    #在使用改进的损失函数，用到了全部的x作为损失函数的参数时，不能使用批量训练 或者批量训练时让对应的x也批量输入 做一个placeholder
            sess.run(train_step,
                feed_dict={x: feature,xjensen:TJensenpower, y: output})
            #定期输出
            if i % 5 ==0:
                #计算损失函数
                train_mse = sess.run(
                    loss_mse, feed_dict={x: feature,xjensen:TJensenpower, y: output})
                test_mse = sess.run(
                    loss_mse, feed_dict={x: test_feature,xjensen:TeJensenpower, y: test_output})
                trainmse.append(train_mse)
                testmse.append(test_mse)
    #            print("After %d  training step(s), trainmse on all data is %g" 
    #                %(i, train_mse))
    #            print("testmse on all data is %g" 
    #                %(test_mse))
    
    ###########################测试准确率#####################################
    #MAE MSE RMSE 校正R平方
        y_predict=sess.run(y_pre,feed_dict={x: test_feature,xjensen:TeJensenpower}) 
        y_true=np.array(test_output)
        y_pred=np.array(y_predict)
        zero=np.zeros(shape=(len(y_true),1))
        print("权重网络网络预测结果")
        y1=y_pred
        a=[]
        #RMSE
        print(np.sqrt(metrics.mean_squared_error(y_true, y_pred))) # 2.847304489713536
        a.append(np.sqrt(metrics.mean_squared_error(y_true, y_pred)))
        # MAE
        print(metrics.mean_absolute_error(y_true, y_pred)) # 1.9285714285714286
        a.append(metrics.mean_absolute_error(y_true, y_pred))
        # MRE
        print(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero)) # 1.9285714285714286
        a.append(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero))
        #r2
        print(metrics.r2_score(y_true,y_pred))
        a.append(metrics.r2_score(y_true,y_pred))
        ########################泛化性################################################
        y_true=np.array(vaild_output)
        y_predict=sess.run(y_pre,feed_dict={x: vaild_feature,xjensen:VJensenpower}) 
        y_pred=np.array(y_predict)
        zero=np.zeros(shape=(len(y_true),1))
        print("权重网络泛化性结果")
        y2=y_pred
        b=[]
        #RMSE
        print(np.sqrt(metrics.mean_squared_error(y_true, y_pred))) # 2.847304489713536
        b.append(np.sqrt(metrics.mean_squared_error(y_true, y_pred)))
        # MAE
        print(metrics.mean_absolute_error(y_true, y_pred)) # 1.9285714285714286
        b.append(metrics.mean_absolute_error(y_true, y_pred))
        # MRE
        print(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero)) # 1.9285714285714286
        b.append(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero))
        #r2
        print(metrics.r2_score(y_true,y_pred))
        b.append(metrics.r2_score(y_true,y_pred))
    return a,y1,b,y2

#########################迁移学习类############################################
def tradaboost(Tf,To,Tef,Teo,Vf,Vo,noise):
    from sklearn.tree import DecisionTreeRegressor
    from sklearn.ensemble import AdaBoostRegressor
    from TwoStageTrAdaBoostR2 import TwoStageTrAdaBoostR2 # import the two-stage algorithm
    rng = np.random.RandomState(1)
    #构造源数据 仿真数据
    phy=pd.read_excel(r"E:/柳靖研究生期间工作/期刊相关文件/最终使用数据\仿真数据.xlsx")
    phy=phy.values.astype(np.float32)
    u1=phy[:,0:6]
    u11=u1[:,0:1]
    noise  = np.random.normal(0,noise,u11.shape)
    y1=phy[:,6:7]+noise
    y1=np.ravel(y1)+rng.normal(0, 0, u1.shape[0])


    #后30%测试
    x_data_test=Tef
    y_data_test=Teo
    y_data_test=np.ravel(y_data_test)+rng.normal(0, 0, x_data_test.shape[0])
    #70%训练
    x_data=Tf
    y_data=To
    y_data=np.ravel(y_data)+rng.normal(0, 0, x_data.shape[0])

    #对目标训练数据进行迁移回归
    len1=len(y_data)
    len2=len(y1)
    X = np.concatenate((u1, x_data))
    y = np.concatenate((y1, y_data))
    sample_size = [len2, len1]
    
    #==============================================================================
    
    n_estimators = 100 #20-30
    steps = 10
    fold = 10
    random_state = np.random.RandomState(1)
    
    #==============================================================================
    #==============================================================================
    
    # 4.2 TwoStageAdaBoostR2
    regr_1 = TwoStageTrAdaBoostR2(DecisionTreeRegressor(max_depth=6),
                          n_estimators = n_estimators, sample_size = sample_size, 
                          steps = steps, fold = fold, 
                          random_state = random_state)
    regr_1.fit(X, y)
    y_pred1 = regr_1.predict(x_data_test)
    #
    # 4.3 As comparision, use AdaBoostR2 without transfer learning
    # 4.3 作为比较，使用无迁移的AdaBoostR2
    #==============================================================================
    regr_2 = AdaBoostRegressor(DecisionTreeRegressor(max_depth=6),
                              n_estimators = n_estimators)
    #==============================================================================
    regr_2.fit(x_data, y_data)
    y_pred2 = regr_2.predict(x_data_test)
    
    

    # 4.5 Calculate mse
    # 4.5 计算mse
    mae_twostageboost = mean_absolute_error(y_data_test, y_pred1)   
    mae_adaboost = mean_absolute_error(y_data_test, y_pred2)
    print("MAE of regular AdaboostR2:", mae_adaboost)
    print("MAE of TwoStageTrAdaboostR2:", mae_twostageboost)
    #===========================Tradaboost预测结果===================================================
    y_true=np.array(y_data_test)
    y_pred=np.array(y_pred1)
    b = y_pred.reshape(len(x_data_test),1)
    if mae_twostageboost<mae_adaboost:
        zero=np.zeros(shape=(len(y_true),1))
        print("Tradaboost算法预测结果")
        y1=b
        c=[]
        #RMSE
        print(np.sqrt(metrics.mean_squared_error(y_true, y_pred))) # 2.847304489713536
        c.append(np.sqrt(metrics.mean_squared_error(y_true, y_pred)))
        # MAE
        print(metrics.mean_absolute_error(y_true, y_pred)) # 1.9285714285714286
        c.append(metrics.mean_absolute_error(y_true, y_pred))
        # MRE
        print(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero)) # 1.9285714285714286
        c.append(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero))
        #r2
        print(metrics.r2_score(y_true,y_pred))
        c.append(metrics.r2_score(y_true,y_pred))
    #===========================adaboost预测结果===================================================
    y_true=np.array(y_data_test)
    y_pred=np.array(y_pred2)
    b = y_pred.reshape(len(x_data_test),1)
    if mae_twostageboost>mae_adaboost:
        zero=np.zeros(shape=(len(y_true),1))
        print("adaboost算法预测结果")
        y1=b
        c=[]
        #RMSE
        print(np.sqrt(metrics.mean_squared_error(y_true, y_pred))) # 2.847304489713536
        c.append(np.sqrt(metrics.mean_squared_error(y_true, y_pred)))
        # MAE
        print(metrics.mean_absolute_error(y_true, y_pred)) # 1.9285714285714286
        c.append(metrics.mean_absolute_error(y_true, y_pred))
        # MRE
        print(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero)) # 1.9285714285714286
        c.append(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero))
        #r2
        print(metrics.r2_score(y_true,y_pred))
        c.append(metrics.r2_score(y_true,y_pred))

    ########################泛化性################################################
    #########################泛化性文件读取################################################
    x_data_add=Vf
    y_data_add=Vo
    ############################Tradaboost算法泛化性####################################
    y_predict_add=regr_1.predict(x_data_add)
    y_true=np.array(y_data_add)
    y_pred=np.array(y_predict_add)
    b = y_pred.reshape(len(x_data_add),1)
    if mae_twostageboost<mae_adaboost:
        zero=np.zeros(shape=(len(y_true),1))
        print("tradaboost泛化性预测结果")
        y2=b
        d=[]
        #RMSE
        print(np.sqrt(metrics.mean_squared_error(y_true, y_pred))) # 2.847304489713536
        d.append(np.sqrt(metrics.mean_squared_error(y_true, y_pred)))
        # MAE
        print(metrics.mean_absolute_error(y_true, y_pred)) # 1.9285714285714286
        d.append(metrics.mean_absolute_error(y_true, y_pred))
        # MRE
        print(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero)) # 1.9285714285714286
        d.append(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero))
        #r2
        print(metrics.r2_score(y_true,y_pred))
        d.append(metrics.r2_score(y_true,y_pred))
    ############################adaboost算法泛化性####################################
    y_predict_add=regr_2.predict(x_data_add)
    y_true=np.array(y_data_add)
    y_pred=np.array(y_predict_add)
    b = y_pred.reshape(len(x_data_add),1)
    if mae_twostageboost>mae_adaboost:
        zero=np.zeros(shape=(len(y_true),1))
        print("adaboost泛化性预测结果")
        y2=b
        d=[]
        #RMSE
        print(np.sqrt(metrics.mean_squared_error(y_true, y_pred))) # 2.847304489713536
        d.append(np.sqrt(metrics.mean_squared_error(y_true, y_pred)))
        # MAE
        print(metrics.mean_absolute_error(y_true, y_pred)) # 1.9285714285714286
        d.append(metrics.mean_absolute_error(y_true, y_pred))
        # MRE
        print(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero)) # 1.9285714285714286
        d.append(metrics.mean_absolute_error(y_true, y_pred)/metrics.mean_absolute_error(y_true, zero))
        #r2
        print(metrics.r2_score(y_true,y_pred))
        d.append(metrics.r2_score(y_true,y_pred))
    return c,y1,d,y2

def loopmodel(i):  #第i次跑
#首先存真实数据集
    Tf,To,Tef,Teo,Vf,Vo=Data(0.7) #70%数据训练 30%测试
    savet=np.concatenate([Tef,Teo],axis=1)
    savet=pd.DataFrame(savet)
    savet.to_csv(r'E:/周寰育/期刊论文/real/test/%d.csv'%i,index=0)
    savev=np.concatenate([Vf,Vo],axis=1)
    savev=pd.DataFrame(savev)
    savev.to_csv(r'E:/周寰育/期刊论文/real/vaild/%d.csv'%i,index=0)
#开始存模型数据
#模型1
    a,y1,b,y2=D2Jensen(Tf,To,Tef,Teo,Vf,Vo)
    savet=np.concatenate([Tef,Teo,y1],axis=1)
    savet=pd.DataFrame(savet)
    savet.to_csv(r'E:/周寰育/期刊论文/model1/test/%d.csv'%i,index=0)
    savev=np.concatenate([Vf,Vo,y2],axis=1)
    savev=pd.DataFrame(savev)
    savev.to_csv(r'E:/周寰育/期刊论文/model1/vaild/%d.csv'%i,index=0)                                               
    # tg = op.load_workbook(r"E:/周寰育/期刊论文/model1/test/output.xlsx")      	# 应先将excel文件放入到工作目录下 
    # sheet = tg["Sheet1"]
    # for j in range(1, len(a)+1):						
    #     sheet.cell(j , i, a[j-1])
    # tg.save("E:/周寰育/期刊论文/model1/test/output.xlsx")
    # vg = op.load_workbook(r"E:/周寰育/期刊论文/model1/vaild/output.xlsx")    	# 应先将excel文件放入到工作目录下 
    # sheet = vg["Sheet1"]
    # for j in range(1, len(b)+1):						
    #     sheet.cell(j , i, b[j-1])
    # vg.save("E:/周寰育/期刊论文/model1/vaild/output.xlsx")
#模型2
    a,y1,b,y2=network(Tf,To,Tef,Teo,Vf,Vo)
    savet=np.concatenate([Tef,Teo,y1],axis=1)
    savet=pd.DataFrame(savet)
    savet.to_csv(r'E:/周寰育/期刊论文/model2/test/%d.csv'%i,index=0)
    savev=np.concatenate([Vf,Vo,y2],axis=1)
    savev=pd.DataFrame(savev)
    savev.to_csv(r'E:/周寰育/期刊论文/model2/vaild/%d.csv'%i,index=0)                                               
    # tg = op.load_workbook(r"E:/周寰育/期刊论文/model2/test/output.xlsx")      	# 应先将excel文件放入到工作目录下 
    # sheet = tg["Sheet1"]
    # for j in range(1, len(a)+1):						
    #     sheet.cell(j , i, a[j-1])
    # tg.save("E:/周寰育/期刊论文/model2/test/output.xlsx")
    # vg = op.load_workbook(r"E:/周寰育/期刊论文/model2/vaild/output.xlsx")    	# 应先将excel文件放入到工作目录下 
    # sheet = vg["Sheet1"]
    # for j in range(1, len(b)+1):						
    #     sheet.cell(j , i, b[j-1])
    # vg.save("E:/周寰育/期刊论文/model2/vaild/output.xlsx")
#模型3
    a,y1,b,y2=Jensennetwork(Tf,To,Tef,Teo,Vf,Vo)
    savet=np.concatenate([Tef,Teo,y1],axis=1)
    savet=pd.DataFrame(savet)
    savet.to_csv(r'E:/周寰育/期刊论文/model3/test/%d.csv'%i,index=0)
    savev=np.concatenate([Vf,Vo,y2],axis=1)
    savev=pd.DataFrame(savev)
    savev.to_csv(r'E:/周寰育/期刊论文/model3/vaild/%d.csv'%i,index=0)                                               
    # tg = op.load_workbook(r"E:/周寰育/期刊论文/model3/test/output.xlsx")      	# 应先将excel文件放入到工作目录下 
    # sheet = tg["Sheet1"]
    # for j in range(1, len(a)+1):						
    #     sheet.cell(j , i, a[j-1])
    # tg.save("E:/周寰育/期刊论文/model3/test/output.xlsx")
    # vg = op.load_workbook(r"E:/周寰育/期刊论文/model3/vaild/output.xlsx")    	# 应先将excel文件放入到工作目录下 
    # sheet = vg["Sheet1"]
    # for j in range(1, len(b)+1):						
    #     sheet.cell(j , i, b[j-1])
    # vg.save("E:/周寰育/期刊论文/model3/vaild/output.xlsx")
#模型4
    a,y1,b,y2=networkJensen(Tf,To,Tef,Teo,Vf,Vo)
    savet=np.concatenate([Tef,Teo,y1],axis=1)
    savet=pd.DataFrame(savet)
    savet.to_csv(r'E:/周寰育/期刊论文/model4/test/%d.csv'%i,index=0)
    savev=np.concatenate([Vf,Vo,y2],axis=1)
    savev=pd.DataFrame(savev)
    savev.to_csv(r'E:/周寰育/期刊论文/model4/vaild/%d.csv'%i,index=0)                                               
    # tg = op.load_workbook(r"E:/周寰育/期刊论文/model4/test/output.xlsx")      	# 应先将excel文件放入到工作目录下 
    # sheet = tg["Sheet1"]
    # for j in range(1, len(a)+1):						
    #     sheet.cell(j , i, a[j-1])
    # tg.save("E:/周寰育/期刊论文/model4/test/output.xlsx")
    # vg = op.load_workbook(r"E:/周寰育/期刊论文/model4/vaild/output.xlsx")    	# 应先将excel文件放入到工作目录下 
    # sheet = vg["Sheet1"]
    # for j in range(1, len(b)+1):						
    #     sheet.cell(j , i, b[j-1])
    # vg.save("E:/周寰育/期刊论文/model4/vaild/output.xlsx")
#模型5
    a,y1,b,y2=weightnetwork(Tf,To,Tef,Teo,Vf,Vo)
    savet=np.concatenate([Tef,Teo,y1],axis=1)
    savet=pd.DataFrame(savet)
    savet.to_csv(r'E:/周寰育/期刊论文/model5/test/%d.csv'%i,index=0)
    savev=np.concatenate([Vf,Vo,y2],axis=1)
    savev=pd.DataFrame(savev)
    savev.to_csv(r'E:/周寰育/期刊论文/model5/vaild/%d.csv'%i,index=0)                                               
    # tg = op.load_workbook(r"E:/周寰育/期刊论文/model5/test/output.xlsx")      	# 应先将excel文件放入到工作目录下 
    # sheet = tg["Sheet1"]
    # for j in range(1, len(a)+1):						
    #     sheet.cell(j , i, a[j-1])
    # tg.save("E:/周寰育/期刊论文/model5/test/output.xlsx")
    # vg = op.load_workbook(r"E:/周寰育/期刊论文/model5/vaild/output.xlsx")    	# 应先将excel文件放入到工作目录下 
    # sheet = vg["Sheet1"]
    # for j in range(1, len(b)+1):						
    #     sheet.cell(j , i, b[j-1])
    # vg.save("E:/周寰育/期刊论文/model5/vaild/output.xlsx")
#模型6
    a,y1,b,y2=tradaboost(Tf,To,Tef,Teo,Vf,Vo,20)
    savet=np.concatenate([Tef,Teo,y1],axis=1)
    savet=pd.DataFrame(savet)
    savet.to_csv(r'E:/周寰育/期刊论文/model6/test/%d.csv'%i,index=0)
    savev=np.concatenate([Vf,Vo,y2],axis=1)
    savev=pd.DataFrame(savev)
    savev.to_csv(r'E:/周寰育/期刊论文/model6/vaild/%d.csv'%i,index=0)                                               
    # tg = op.load_workbook(r"E:/周寰育/期刊论文/model6/test/output.xlsx")      	# 应先将excel文件放入到工作目录下 
    # sheet = tg["Sheet1"]
    # for j in range(1, len(a)+1):						
    #     sheet.cell(j , i, a[j-1])
    # tg.save("E:/周寰育/期刊论文/model6/test/output.xlsx")
    # vg = op.load_workbook(r"E:/周寰育/期刊论文/model6/vaild/output.xlsx")    	# 应先将excel文件放入到工作目录下 
    # sheet = vg["Sheet1"]
    # for j in range(1, len(b)+1):						
    #     sheet.cell(j , i, b[j-1])
    # vg.save("E:/周寰育/期刊论文/model6/vaild/output.xlsx")


#for i in range(3,4):
#    loopmodel(i)
loopmodel(6)

